from InstagramAPI import InstagramAPI
import time
from openpyxl import Workbook
import datetime
from openpyxl import load_workbook

class UserDetails:

    def getTotalFollowers(self, credentials, filePath):
        for key, value in credentials.items():
            user_id =  value.username_id
            followers = []
            next_max_id = True
            todayDate = datetime.datetime.now()
            fileName = key+'-'+'Followers for last 24 hrs-'+todayDate.strftime('%b %d,%Y')+'.xlsx'
            workbook = Workbook()
            worksheet = workbook.active
            while next_max_id:
                # first iteration hack
                if next_max_id is True:
                    next_max_id = ''

                _ = value.getUserFollowers(user_id, maxid=next_max_id)
                followers.extend(value.LastJson.get('users', []))
                next_max_id = value.LastJson.get('next_max_id', '')

            count = 1
            for follower in followers:
                worksheet.cell(row=count, column=1, value=follower['username'])
                count = count + 1
            workbook.save(filePath + fileName)