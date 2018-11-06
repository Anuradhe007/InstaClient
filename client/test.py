import datetime
import sched, time

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

# s = sched.scheduler(time.time, time.sleep)
# def print_time(a='default'):
#      print("From print_time", time.time(), a)
#
# def print_some_times():
#      print(time.time())
#      s.enter(10, 1, print_time)
#      s.enter(5, 2, print_time, argument=('positional',))
#      s.enter(5, 1, print_time, kwargs={'a': 'keyword'})
#      s.run()
#      print(time.time())
#
# print_some_times()

#readable = datetime.datetime.fromtimestamp(1540638110).isoformat()
readable = datetime.datetime.fromtimestamp(1540638110)
current = datetime.datetime.now() - datetime.timedelta(minutes=15)
#print(readable<current)